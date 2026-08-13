# -*- coding: utf-8 -*-
"""生成 LED 屏询价指引 xlsx（Step 2/3 输出：询价总览/价格汇总/待澄清与风险/询价策略/厂商报价对比/分项询价单）
用法: python3 gen_inquiry_xlsx.py <输出路径.xlsx>
样式遵循 SKILL 规则 14: 表头深蓝 2F5496 白字加粗居中、数据行微软雅黑10细边框自动换行、
合计行浅蓝 D6E4F0 加粗、列宽按类型、行高按文本长度估算
价格数据来源: 2026-08-11 三份厂商报价实价（洲明/利亚德/海康项目单）+ AI 估算
"""
import math
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10, color="000000")
GROUP_FILL = PatternFill("solid", fgColor="D6E4F0")
GROUP_FONT = Font(name="微软雅黑", size=10, bold=True, color="2F5496")
TOTAL_FONT = Font(name="微软雅黑", size=10, bold=True, color="000000")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
NOTE_FONT = Font(name="微软雅黑", size=10, bold=True, color="C00000")


def est_lines(text, width):
    """估算单元格文本所需行数: 全角字符按 2 个宽度单位计"""
    units = 0
    for line in str(text).split("\n"):
        units = max(units, sum(2 if ord(ch) > 127 else 1 for ch in line))
    return max(1, math.ceil(units / max(width - 2, 4)))


def write_sheet(ws, headers, rows, widths, centers=None, group_idx=None, total_idx=None,
                extra=None, header_row=1):
    """写入带表头的表格并美化。
    centers: 居中的列下标集合(0-based); group_idx/total_idx: 分组/合计行下标;
    extra: (起始行, 文本, 合并宽度) 列表, 写在表格下方; header_row: 表头所在行(1-based)
    """
    centers = centers or set()
    ncols = len(headers)
    r0 = header_row
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, CENTER, BORDER
        ws.column_dimensions[chr(64 + c)].width = w
    ws.row_dimensions[r0].height = 22
    for i, row in enumerate(rows):
        r = r0 + 1 + i
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font, cell.border = DATA_FONT, BORDER
            cell.alignment = CENTER if (c - 1) in centers else LEFT
            if group_idx and i in group_idx:
                cell.fill = GROUP_FILL
                cell.font = GROUP_FONT
            if total_idx and i in total_idx:
                cell.fill = GROUP_FILL
                cell.font = TOTAL_FONT
            ws.row_dimensions[r].height = max(
                est_lines(v, widths[c - 1]) * 14.5, ws.row_dimensions[r].height or 0)
    if extra:
        start = r0 + len(rows) + 2
        for i, (text, n_merge) in enumerate(extra):
            r = start + i
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_merge or ncols)
            cell = ws.cell(row=r, column=1, value=text)
            cell.font, cell.alignment = NOTE_FONT, LEFT
            ws.row_dimensions[r].height = est_lines(text, sum(widths[:n_merge or ncols])) * 14.5


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "LED屏询价指引.xlsx"
    wb = Workbook()

    # Sheet 1 询价总览
    ws = wb.active
    ws.title = "询价总览"
    write_sheet(ws,
        ["#", "设备名称", "推荐询价厂商", "参考型号", "预估整体打包价", "数量", "预估小计", "置信度", "询价优先级", "关键假设"],
        [[1,
          "室内LED全彩屏 P1.25 COB\n(3.6×2.08m ≈ 7.49㎡)",
          "利亚德 / 洲明 / 艾比森 / 强力巨彩",
          "利亚德 VCS012、洲明 UMiniP1.2（同规格实价已确认）",
          "¥210,000-310,000", "1 套", "¥210,000-310,000", "🟢 实价修正", "🔴立即",
          "按 7.49㎡ 小屏较 34㎡ 报价上浮 10-20% 估；已含处理器/配电柜/支架/包边/安装/措施费全部费用（含13%税）"]],
        [5, 30, 28, 40, 16, 8, 16, 12, 10, 46],
        centers={0, 4, 5, 6, 7, 8})

    # Sheet 2 价格汇总（实价修正）
    ws = wb.create_sheet("价格汇总")
    write_sheet(ws,
        ["分项", "修正后参考区间", "数量/口径", "价格依据（实价）", "说明"],
        [["COB 屏体（模组+压铸铝箱体）", "¥21,000-26,000/㎡（本次7.49㎡）",
          "7.49㎡", "洲明 UMiniP1.2 ¥19,000/㎡；利亚德 VCS012 ¥21,650/㎡（均34㎡规模 COB倒装）",
          "小屏单价上浮 10-20%；SMD 工艺仅 ¥10,900/㎡（海康项目单，不满足 COB 要求）"],
         ["拼接处理器（拼控整机）", "¥20,000-40,000/台", "1台",
          "海康 DS-B30-S11（10槽位机架式）¥35,960",
          "若采用分布式架构可替代：节点 ¥2,900/台（洲明实价）"],
         ["输入卡（4路HDMI 1080p）", "¥7,000-10,000/卡", "按路数",
          "海康 DS-B30-04HI ¥8,325",
          "编码 4路1080p60"],
         ["输入卡 4K（2路DP）", "¥14,000-18,000/卡", "按路数",
          "海康 DS-B30-02DPI/4K ¥16,980",
          "2路4K60 输入"],
         ["输出卡（4路HDMI）", "¥9,000-12,000/卡", "按路数",
          "海康 DS-B30-04HO ¥10,500",
          "4路输出+64路解码通道"],
         ["LED发送卡/发送盒", "¥2,300-2,800/台", "按屏",
          "利亚德 SM600 ¥2,300；卡莱特 ¥2,800",
          "4网口带载230万点"],
         ["配电柜", "¥800-5,000/台", "1台",
          "利亚德 40KW ¥800；海康 50KW 智能（远程/逐路上电）¥4,800",
          "本项目按 7.49㎡ 需求 40KW 档即可，¥800-2,000"],
         ["支架/钢骨架", "¥700-1,200/㎡", "7.49㎡",
          "海康一体化支架 ¥800/㎡（87.3㎡，全封闭防尘）",
          "小屏单价上浮；异形墙/轻质隔墙另计"],
         ["包边", "¥9,000-23,000", "周长11.4m",
          "🟡 AI 估算", "不锈钢/铝单板 ¥800-2,000/延米"],
         ["线材辅材", "¥100-400/㎡", "7.49㎡",
          "利亚德 ¥400/㎡；海康辅材包 ¥8,000/项",
          "含电源线/网线/排线"],
         ["安装调试", "¥1,500-3,000/㎡", "7.49㎡",
          "🟡 AI 估算", "吊装+逐点校正+联调+运输上楼"],
         ["措施费/成品保护", "总价 3-5%", "—",
          "🟡 AI 估算", "含二次深化设计、验收配合"],
         ["合计（整体打包）", "¥210,000-310,000", "1套",
          "三家实价 + AI 估算", "≈¥28,000-41,000/㎡ 含全部费用"]],
        [24, 26, 12, 44, 38],
        centers={2},
        total_idx={12})

    # Sheet 3 待澄清事项与风险（报价回填后更新）
    ws = wb.create_sheet("待澄清事项")
    write_sheet(ws,
        ["#", "事项", "影响", "处理建议"],
        [[1, "平均功耗：两家报价按塔联矿参数承诺平均≤162W/㎡，本次招标要求≤130W/㎡", "🔻 负偏离风险，162>130 可致投标无效",
          "询价时要求厂商按 130W/㎡ 重新核对功耗设计（共阳驱动+节能电源）"],
         [2, "▲光生物安全（LB≤5W·m⁻²·sr⁻¹）+ CMA 报告：两家报价未应答", "投标资格项，无报告=废标",
          "要求随报价附 CMA 检测报告，报告抬头与投标主体一致"],
         [3, "▲防透光设计 + CMA 报告：两家报价未应答", "投标资格项，无报告=废标", "同上，两份报告都要"],
         [4, "BJ≥20 鉴别等级：报价未列", "无法判断", "要求厂商提供鉴别等级检测报告"],
         [5, "信噪比≥60dB/延迟≤500ns：报价仅写“延时1帧”无量化值", "无法判断", "要求厂商补充量化参数"],
         [6, "线性/三角形像素排列：报价未列", "无法判断", "确认是否支持"],
         [7, "SMD 工艺 ¥10,900/㎡（海康项目单）不满足 COB 要求", "不可用但提示成本差异", "若甲方允许同规格 SMD+覆膜方案，成本降近一半，可作为备选谈判策略"],
         [8, "安装面承重/包边材质需现场勘察", "异形墙推高支架费", "先勘察，二次深化设计以实测为准"]],
        [5, 40, 32, 40],
        centers={0})

    # Sheet 4 询价策略建议（报价回填后更新）
    ws = wb.create_sheet("询价策略建议")
    write_sheet(ws,
        ["优先级", "行动", "厂商数量", "比价要点"],
        [["🔴立即", "向 4 家厂商发送分项询价单（按 130W/㎡ 功耗要求）", "4 家（利亚德/洲明/艾比森/强力巨彩）",
          "要求分项列价勿打包；确认 COB 原厂产线；CMA 报告×2 随报价附上"],
         ["🟡本周", "回收报价后进入 Step 3 参数比对", "—",
          "重点核对：平均功耗≤130W/㎡、亮度600/均匀性98%/刷新率3840Hz/视角160°/CMA 报告"],
         ["🟢可缓", "确认成交价后沉淀 equipment-db", "—",
          "写入 COB ¥19,000-21,650/㎡、支架 ¥800/㎡、板卡 ¥8,325-16,980、拼控 ¥35,960 等实价"],
         ["📊已确认", "三份报价横向对比（详见“厂商报价对比”Sheet）", "洲明/利亚德/海康项目",
          "COB 实价 ¥19,000-21,650/㎡；SMD 仅 ¥10,900/㎡；支架 ¥800/㎡；拼控 DS-B30-S11 ¥35,960"]],
        [10, 40, 30, 54],
        centers={0})

    # Sheet 5 厂商报价对比（内部参考）
    ws = wb.create_sheet("厂商报价对比")
    write_sheet(ws,
        ["设备", "洲明（塔联矿）", "利亚德（塔联矿）", "海康项目单", "参考区间（本次用）"],
        [["LED屏体 P1.25", "UMiniP1.2 COB ¥19,000/㎡", "VCS012 COB ¥21,650/㎡",
          "GLW1.2 SMD ¥10,900/㎡", "COB ¥21,000-26,000/㎡（7.49㎡小屏）"],
         ["发送卡", "SD700E 未报", "SM600 ¥2,300/个", "卡莱特 ¥2,800/台", "¥2,300-2,800/台"],
         ["拼接处理器", "分布式替代（见下）", "未单列", "DS-B30-S11 ¥35,960/台", "¥20,000-40,000/台"],
         ["输入卡", "—", "—", "DS-B30-04HI ¥8,325/卡", "¥7,000-10,000/卡"],
         ["输入卡4K", "—", "—", "DS-B30-02DPI/4K ¥16,980/卡", "¥14,000-18,000/卡"],
         ["输出卡", "—", "—", "DS-B30-04HO ¥10,500/卡", "¥9,000-12,000/卡"],
         ["分布式节点(替代拼控)", "输入/输出/KVM ¥2,900/台", "—", "—", "¥2,900/台"],
         ["配电柜", "未单列", "40KW ¥800/个", "50KW智能 ¥4,800/台", "¥800-5,000/台"],
         ["支架/钢骨架", "未单列", "未报价(0)", "一体化 ¥800/㎡", "¥700-1,200/㎡"],
         ["线材辅材", "未单列", "¥400/㎡", "¥8,000/项", "¥100-400/㎡"],
         ["中控主机", "¥18,000/台", "—", "DS-B86CS ¥4,350/台", "¥4,000-18,000/台"],
         ["管控平台软件", "¥32,800/套", "—", "—", "¥20,000-35,000/套"],
         ["合计口径", "¥826,480（34㎡含分布式系统）", "¥787,336（34㎡不含钢架安装）", "¥1,797,121（72.9㎡含音频/工作站）", "—"]],
        [20, 32, 32, 30, 36],
        centers={0},
        total_idx={12},
        extra=[("注：三份报价均为塔联矿/海康项目实际报价（2026-08 取得）；本项目 7.49㎡ COB 屏按小屏上浮 10-20% 修正。", 5)])

    # Sheet 6 分项询价单（外部用）
    ws = wb.create_sheet("LED屏询价单(外部)")
    write_sheet(ws,
        ["#", "设备名称", "规格参数", "数量", "单位", "品牌范围",
         "请填写：报价型号", "请填写：单价(含税)", "请填写：交货周期", "请填写：质保期", "备注"],
        [[1, "COB屏体 P1.25",
          "3.6×2.08m≈7.49㎡；点距≤1.25mm、COB倒装封装；亮度≥600cd/㎡、均匀性≥98%、BJ≥20；对比度≥10000:1；视角≥160°；刷新率≥3840Hz；功耗峰值≤380W/㎡、平均≤130W/㎡；压铸铝箱体、无风扇；反光率≤1%",
          "7.49", "㎡", "COB一线品牌", "", "", "", "",
          "须注明 COB 封装是否原厂产线（非二次贴装）；附 CMA 检测报告（光生物安全、防透光）；平均功耗须≤130W/㎡"],
         [2, "拼接处理器", "支持 4K60 输入、≥2×2 拼接输出、信噪比≥60dB、延迟≤500ns", "1", "台",
          "诺瓦/卡莱特/海康/原厂", "", "", "", "", "通道数≥2入4出；或分布式节点方案"],
         [3, "配电柜", "智能配电，逐路上电、远程控制；40KW 档", "1", "套", "—", "", "", "", "",
          "含远程控制；按屏体功耗选型"],
         [4, "钢骨架+包边", "国标 Q235B 钢材 + 不锈钢/铝单板包边（与装修协调）", "1", "项", "—", "", "", "", "",
          "含二次深化设计；尺寸以实测为准"],
         [5, "安装调试+措施费", "吊装/脚手架 + 逐点校正 + 系统联调 + 运输上楼 + 成品保护", "1", "项", "—", "", "", "", "",
          "含运输上楼、验收配合"]],
        [5, 18, 56, 8, 7, 14, 16, 14, 12, 12, 30],
        centers={0, 3, 4, 6, 7, 8, 9},
        extra=[("★ 请按上述分项报价，勿合并打包（便于比价）；★ 附 2 份 CMA 检测报告（光生物安全+防透光）；★ 注明 COB 封装是否为原厂产线；★ 平均功耗须承诺≤130W/㎡。", 11)])
    # 表格上方插入供应商信息行
    ws.insert_rows(1, 3)
    meta = [("供应商名称：", ""), ("询价日期：", ""), ("有效期至：", "")]
    for i, (label, _) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=i, column=1).alignment = LEFT

    wb.save(out)
    print("已生成:", out)
    for s in wb.sheetnames:
        print("  Sheet:", s)


if __name__ == "__main__":
    main()
